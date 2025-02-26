import kivy
from kivy.app import App
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.gridlayout import GridLayout
from kivy.uix.label import Label
from kivy.uix.textinput import TextInput
from kivy.uix.button import Button
from kivy.uix.image import Image
from kivy.uix.scrollview import ScrollView
from kivy.core.window import Window
from kivy.utils import platform
from kivy.properties import StringProperty
from kivy.uix.popup import Popup
from kivy.clock import mainthread
from datetime import datetime
import os
import subprocess
import openpyxl
from kivy.uix.widget import Widget

kivy.require('2.0.0')


class LoginApp(App):
    current_path = StringProperty("Data")
    username = StringProperty("")
    password = StringProperty("")
    welcome_username = StringProperty("")
    app_title = "Learning Application Demo"  # Title

    def build(self):
        Window.clearcolor = (248 / 255, 248 / 255, 248 / 255, 1)  # background color
        self.root_layout = BoxLayout(orientation='vertical')
        self.root_layout.add_widget(self.login_screen())
        return self.root_layout

    def create_top_layout(self):
        top_layout = BoxLayout(orientation='horizontal', size_hint_y=None, height=150)
        logo = Image(source='backend/Logo.jpg', size_hint=(None, None), size=(150, 150))
        top_layout.add_widget(logo)

        # Title Label in the Middle
        title_label = Label(text=self.app_title, color=(51 / 255, 51 / 255, 51 / 255, 1), font_size=24,
                            halign='center', valign='middle', size_hint_x=0.6)
        top_layout.add_widget(title_label)

        # Spacer to push content to the right
        top_layout.add_widget(Widget())
        return top_layout

    def login_screen(self):
        layout = BoxLayout(orientation='vertical', padding=20, spacing=20)

        top_layout = self.create_top_layout()
        layout.add_widget(top_layout)

        # Centering layout
        center_layout = BoxLayout(orientation='vertical', size_hint_y=0.5,
                                  pos_hint={'center_x': 0.5, 'center_y': 0.5})

        form_layout = GridLayout(cols=2, spacing=20, size_hint_y=None, height=150)
        form_layout.bind(minimum_height=form_layout.setter('height'))

        form_layout.add_widget(Label(text="Username:", color=(51 / 255, 51 / 255, 51 / 255, 1), font_size=18))
        self.username_input = TextInput(multiline=False, size_hint_y=None, height=50, font_size=18)
        form_layout.add_widget(self.username_input)

        form_layout.add_widget(Label(text="Password:", color=(51 / 255, 51 / 255, 51 / 255, 1), font_size=18))
        self.password_input = TextInput(multiline=False, size_hint_y=None, height=50, password=True,
                                         font_size=18)
        form_layout.add_widget(self.password_input)

        center_layout.add_widget(form_layout)

        # Add a spacer to increase spacing between form and button
        center_layout.add_widget(Widget(size_hint_y=0.1))

        login_button = Button(text="Login", on_press=self.check_credentials, size_hint_y=None, height=50,
                              background_color=(102 / 255, 178 / 255, 1, 1), font_size=18)
        center_layout.add_widget(login_button)

        layout.add_widget(center_layout)
        layout.add_widget(Widget(size_hint_y=1))  # Add a spacer to push content to the top

        return layout

    def welcome_screen(self):
        layout = BoxLayout(orientation='vertical', padding=20, spacing=10)

        top_layout = self.create_top_layout()
        layout.add_widget(top_layout)

        bottom_layout = BoxLayout(orientation='horizontal', size_hint_y=None, height=75)
        # Spacer to push logout button to the right
        bottom_layout.add_widget(Widget())

        logout_button = Button(text="Logout", on_press=self.logout, size_hint_x=None, width=100, height=75,
                               background_color=(102 / 255, 178 / 255, 1, 1))
        bottom_layout.add_widget(logout_button)

        layout.add_widget(bottom_layout)

        content_layout = BoxLayout(orientation='vertical', size_hint=(1, 0.5),
                                   pos_hint={'center_x': 0.5, 'center_y': 0.5})
        content_layout.add_widget(
            Label(text=f"Welcome {self.username}!", font_size='24sp', color=(0, 66 / 255, 153 / 255, 1)))
        content_layout.add_widget(
            Label(text="You are logged in successfully!", color=(51 / 255, 51 / 255, 51 / 255, 1)))
        continue_button = Button(text="Continue to Learning App", on_press=self.to_learning_screen,
                                 size_hint_y=None,
                                 height=40, background_color=(102 / 255, 178 / 255, 1, 1))
        content_layout.add_widget(continue_button)
        layout.add_widget(content_layout)

        return layout

    def learning_screen(self):
        layout = BoxLayout(orientation='vertical', padding=20, spacing=10)

        top_layout = self.create_top_layout()
        layout.add_widget(top_layout)

        bottom_layout = BoxLayout(orientation='horizontal', size_hint_y=None, height=75)
        back_button = Button(text="Back", on_press=self.go_back, size_hint_x=0.1, height=75,
                             background_color=(102 / 255, 178 / 255, 1, 1))
        self.back_button = back_button
        self.back_button.disabled = (self.current_path == "Data")  # Disable button if in "Data" folder
        bottom_layout.add_widget(back_button)
        path_label = Label(text=f"Current Path: {self.current_path}", size_hint_x=0.4, height=75)
        self.path_label = path_label
        bottom_layout.add_widget(path_label)

        # Spacer to push logout button to the right
        bottom_layout.add_widget(Widget())

        logout_button = Button(text="Logout", on_press=self.logout, size_hint_x=0.2, height=75,
                               background_color=(102 / 255, 178 / 255, 1, 1))
        bottom_layout.add_widget(logout_button)

        layout.add_widget(bottom_layout)

        scroll_view = ScrollView()
        self.content_layout = BoxLayout(orientation='vertical', size_hint_y=None)
        self.content_layout.bind(minimum_height=self.content_layout.setter('height'))
        scroll_view.add_widget(self.content_layout)
        layout.add_widget(scroll_view)
        self.display_folder_contents()
        return layout

    def display_folder_contents(self):
        self.content_layout.clear_widgets()
        try:
            items = os.listdir(self.current_path)
            for item in items:
                item_path = os.path.join(self.current_path, item)
                if os.path.isfile(item_path):
                    button = Button(text=item,
                                    on_press=lambda instance, path=item_path: self.open_file(path),
                                    size_hint_y=None,
                                    height=40, background_color=(102 / 255, 178 / 255, 1, 1))
                    self.content_layout.add_widget(button)
                elif os.path.isdir(item_path):
                    button = Button(text=item,
                                    on_press=lambda instance, path=item_path: self.navigate_to_folder(path),
                                    size_hint_y=None, height=40, background_color=(102 / 255, 178 / 255, 1, 1))
                    self.content_layout.add_widget(button)
        except Exception as e:
            self.show_popup("Error", f"Error displaying folder contents: {e}")

    def open_file(self, file_path):
        try:
            if platform == 'win':
                os.startfile(file_path)
            elif platform == 'linux' or platform == 'macosx':
                subprocess.call(('open', file_path))
            elif platform == 'android':
                subprocess.call(('xdg-open', file_path))  # try this or find android specific method.
        except Exception as e:
            self.show_popup("Error", f"Error opening file: {e}")

    def navigate_to_folder(self, folder_path):
        self.current_path = folder_path
        self.path_label.text = f"Current Path: {self.current_path}"
        self.back_button.disabled = (self.current_path == "Data")
        self.display_folder_contents()

    def go_back(self, instance):
        parent_path = os.path.dirname(self.current_path)
        if parent_path != self.current_path:
            self.current_path = parent_path
            self.path_label.text = f"Current Path: {self.current_path}"
            self.back_button.disabled = (self.current_path == "Data")
            self.display_folder_contents()

    def read_credentials_from_excel(self, file_path):
        try:
            wb = openpyxl.load_workbook(file_path)
            sheet = wb.active
            credentials = {}

            for row in range(2, sheet.max_row + 1):  # Assuming header is in row 1
                username = str(sheet.cell(row=row,
                                          column=1).value).strip().lower()  # Convert to string, remove leading/trailing spaces, and convert to lowercase
                password = sheet.cell(row=row, column=2).value
                valid_till = sheet.cell(row=row, column=3).value

                if valid_till is not None:
                    if isinstance(valid_till, str):  # Check if valid_till is a string
                        valid_till = datetime.strptime(valid_till,
                                                       '%Y-%m-%d')  # Assuming date format is YYYY-MM-DD
                    elif isinstance(valid_till, datetime):  # If it's already a datetime object, no need to parse
                        pass
                    else:
                        self.show_popup("Error", f"Invalid date format for {username}.")
                        continue

                credentials[username] = (password, valid_till)

            return credentials

        except Exception as e:
            self.show_popup("Error", f"Failed to read credentials: {e}")
            return {}

    def check_credentials(self, instance):
        # Get text from input fields
        self.username = self.username_input.text.strip().lower()
        self.password = self.password_input.text

        # Read credentials from Excel file
        credentials = self.read_credentials_from_excel("Backend/Credentials.xlsx")

        # Validate credentials
        if self.username in credentials:
            stored_password, valid_till = credentials[self.username]
            if self.password == stored_password:
                if valid_till is None or datetime.now() <= valid_till:
                    self.to_welcome_screen()
                else:
                    self.show_popup("Login Failed", "Login credentials expired. Please contact support.")
            else:
                self.show_popup("Login Failed", "Invalid password.")
        else:
            self.show_popup("Error",
                            f"Username '{self.username}' not found in credentials. Please contact support.")

    def to_welcome_screen(self):
        self.root_layout.clear_widgets()
        self.root_layout.add_widget(self.welcome_screen())

    def to_learning_screen(self, instance):
        self.root_layout.clear_widgets()
        self.root_layout.add_widget(self.learning_screen())

    def logout(self, instance):
        self.root_layout.clear_widgets()
        self.root_layout.add_widget(self.login_screen())

    def show_popup(self, title, message):
        popup = Popup(title=title, content=Label(text=message), size_hint=(None, None), size=(400, 100))
        popup.open()


if __name__ == "__main__":
    LoginApp().run()
